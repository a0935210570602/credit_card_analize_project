import pandas as pd
import numpy as np
from openpyxl import load_workbook

def write_to_excel(data,file_name,SheetName,mode=0):
    if mode == 0:
        df = pd.DataFrame(data)
    elif mode == 1:
        title = ["銀行","信用卡","年費(NT)","國外消費現金回饋(%)","國內消費現金回饋(%)",
                "國外哩程回饋(NT/哩)","國內哩程回饋(NT/哩)","加油現金回饋(%)","加油優惠(NT/公升)",
                "周末電影折扣(折)","平日電影折扣(折)","最高分期期數(期)","最低可分期金額(NT)","回饋上限(NT)",
                "網購回饋(%)","首刷送行李箱(1/0)","首刷哩程回饋(哩)","首刷現金回饋(NT)","刷保費哩程回饋(%)",
                "最高分期期數(0利率)(期)","刷保費現金回饋(%)","最低可分期金額(NT)","紅利折抵金","紅利換哩程","最高基點倍數"]

        df = pd.DataFrame(data,columns = title)

    # 'test.xlsx' 為檔案名(要先建立好，並放在同個目錄下)
    #可以用 help(pd.ExcelWriter) 去看該函數的參數
    #此處比較像在搜尋該excel檔是否存在
    writer = pd.ExcelWriter( file_name ,engine='openpyxl')

    #為了不使新的輸入資料覆蓋舊的，
    #所以會先使用【 writer.book = load_workbook('test.xlsx') 】看裡面有沒有東西
    #但是，因為一開始建好的檔案是沒有東西的，所以該行會有錯誤，所以
    #先看裡面有沒有檔案:try ， 若有，繼續執行try，即 將資料輸入excel
    #若沒有，則執行:except ，不檢查裡面有沒有東西，就直接寫入(會覆蓋)
    try:
        writer.book = load_workbook( file_name )
        df.to_excel(writer,sheet_name= SheetName)
        print("try: ")
    except:
        df.to_excel(writer,sheet_name= SheetName)
        print("except: ")

    writer.save()
    writer.close()