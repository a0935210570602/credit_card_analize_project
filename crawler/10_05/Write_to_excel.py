import pandas as pd
import numpy as np
from openpyxl import load_workbook

def write_to_excel(data,file_name,SheetName):

    df = pd.DataFrame(data)

    # 'test.xlsx' 為檔案名(要先建立好，並放在同個目錄下)
    #可以用 help(pd.ExcelWriter) 去看該函數的參數
    #此處比較像在搜尋該excel檔是否存在
    writer = pd.ExcelWriter( file_name ,engine='openpyxl')

    #為了不使新的輸入資料覆蓋舊的，
    #所以會先使用【 writer.book = load_workbook('test.xlsx') 】看裡面有沒有東西
    #但是，因為一開始建好的檔案是沒有東西的，所以該行會有錯誤，所以
    #先看裡面有沒有檔案:try ， 若有，繼續執行try，即 將資料輸入excel
    #若沒有，則執行:except ，不檢查裡面有沒有東西，就直接寫入(會覆蓋)
    try:
        writer.book = load_workbook( file_name )
        df.to_excel(writer,sheet_name= SheetName)
        print("try: ")
    except:
        df.to_excel(writer,sheet_name= SheetName)
        print("except: ")

    writer.save()
    writer.close()